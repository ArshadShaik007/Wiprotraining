import openpyxl

# create object of workbook
wb_object=openpyxl.load_workbook('samplexl.xlsx')

sheet_obj = wb_object.active

cell_obj = sheet_obj.cell(row=1, column=1)

# specific sheet
spec_sheet=sheets_dict['sheet1']

print(cell_obj.value)

# print all row

print(sheet_obj.max_row)

# print all column

print(sheet_obj.max_column)

# print all column names

for i in range (1,sheet_obj.max_column+1):
    cell_obj = sheet_obj.cell(row=1,column=i)
    print(cell_obj.value)

# print 1st column values
for i in range (1,sheet_obj.max_row+1):
    cell_obj = sheet_obj.cell(row=i,column=1)
    print(cell_obj.value)

# get perticular row value

for i in range(1, sheet_obj.max_column + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value)