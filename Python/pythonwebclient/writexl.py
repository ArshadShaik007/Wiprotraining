import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

c1 = sheet.cell(row=1,column=1)
c1.value = "john"

c1= sheet.cell(row=1,column=2)
c1.value = "sunny"

c1= sheet.cell(row=1,column=3)
c1.value = "male"

wb.save('demo.xlsx')